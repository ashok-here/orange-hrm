from selenium import webdriver
from selenium.webdriver import ActionChains
import openpyxl

driver = webdriver.Chrome(executable_path="C:/Users/areddymasi/PycharmProjects/orange-hrm/Drivers/chromedriver.exe")

driver.maximize_window()
driver.get("https://opensource-demo.orangehrmlive.com/")

driver.find_element_by_name("txtUsername").send_keys("Admin")
driver.find_element_by_name("txtPassword").send_keys("admin123")
driver.find_element_by_id("btnLogin").click()

pim = driver.find_element_by_id("menu_pim_viewPimModule")
employee_list = driver.find_element_by_id("menu_pim_viewEmployeeList")
actions = ActionChains(driver)
actions.move_to_element(pim).move_to_element(employee_list).click().perform()

row_count = len(driver.find_elements_by_xpath("//table[@id='resultTable']/tbody/tr"))
print("Rows Count is:", row_count)

col_count = len(driver.find_elements_by_xpath("//table[@id='resultTable']/thead/tr/th"))
print("Columns Count is:", col_count)

path = "D:/Openpyxl/OrangeHRM_EmployeeList_Grid_Data.xlsx"
workbook = openpyxl.Workbook()
workbook.save(path)
worksheet = workbook.active
worksheet.title = "EmployeeList"

for x in range(1, 2):
    for y in range(1, col_count+1):
        col_headers = driver.find_element_by_xpath("//table[@id='resultTable']/thead/tr/th["+str(y)+"]").text
        worksheet.cell(x, y).value = col_headers
        workbook.save(path)

for r in range(1, row_count+1):
    for c in range(1, col_count+1):
        values = driver.find_element_by_xpath("//table[@id='resultTable']/tbody/tr["+str(r)+"]/td["+str(c)+"]").text
        worksheet.cell(r+1, c).value = values
        workbook.save(path)

